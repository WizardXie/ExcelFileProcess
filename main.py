# This is a sample Python script.
# 注意事项：
# 开始合并数据之前务必对原始文档进行备份
# 开始合并前须检查各个文档的格式是否保持一致
# 开始合并前各个文档必须是关闭的状态
# 合并数据时主文档必须保持空白无数据的状态
# 和并完成后最好再检查核对一下数据是否一致
# 只能处理xlsx文件格式
# 所有文档放在当前目录下
# 工作簿中只能有一个工作表
# 没有考虑关键字段值重复的问题
# 进行了部分异常处理，还不够完善

# 完成！ 警察学系的表结构可能有问题，需要把空白的行和列删去才能正常导入

# TODO:

# 如果主文档中用于比较的关键字段没有数据，也应该报错
# 提示将要从n个文档中合并数据，是否继续
# 将参数设置调整为通过配置文件实现，而非每次运行的时候要求用户输入
# 每次合并完成后保存成一个新的文档（文件名与合并时间关联）

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

# SheetName = "Sheet1"
import openpyxl
import os
import sys
import datetime
from shutil import copyfile

ReservedRow = 1  # 工作表中保留的行数 - 通常表示表头
ReservedColumn = 3  # 工作表中保留的列数 - 通常包括序号、姓名、编号等
KeyColumn = 2  # 工作表中用于进行对比的关键字段所在列
fileNameList = []  # 用于保存在当前目录中找到的Excel文档的文件名列表


def make_settings():
    prompt = "\n请设置Excel文档工作表中需要保留的行数、列数，以及进行对比的关键字段所在列："
    error_prompt = "输入有误，请重新输入!"
    while True:
        setreservedrowvalue, setreservedcolumnvalue, setkeycolumnvalue = input(
            prompt).split()
        try:
            setreservedrowvalue = int(setreservedrowvalue)
            setreservedcolumnvalue = int(setreservedcolumnvalue)
            setkeycolumnvalue = int(setkeycolumnvalue)
        except ValueError:
            print(error_prompt)
            continue
        else:
            break
    settings = tuple(
        [setreservedrowvalue, setreservedcolumnvalue, setkeycolumnvalue])
    return settings


def display_currentpath_info():  # 输出提示，显示当前目录中Excel文档的相关信息
    currentPath = os.getcwd()
    print("当前目录是: {0}".format(currentPath))
    allFileList = os.listdir(currentPath)
    for fileName in allFileList:
        if fileName.endswith('.xlsx'):
            fileNameList.append(fileName)
    if len(fileNameList) == 0:
        print("在当前目录中没有检测到EXCEL文档，无法继续合并操作，程序退出。")
        exit(1)
    fileNameList.sort()
    print("在当前目录中检测到{0}个EXCEL文档：".format(len(fileNameList)))
    for index in range(1, len(fileNameList) + 1):
        print("\t{0}:\t{1}".format(str(index), fileNameList[index - 1]))


def get_choice():  # 输出提示，并接受用户的选择（同时进行异常处理），返回对应的数值
    prompt = "\n请选择要存放合并数据的Excel文档的序号（选择0将退出程序）："
    error_prompt = "输入有误，请重新输入!"
    while True:  # 接受用户的输入，并进行异常处理
        choice = input(prompt)
        try:
            choice = int(choice)
        except ValueError:
            print(error_prompt)
            continue
        if choice == 0:
            exit(0)
        if choice in range(1, len(fileNameList) + 1):
            break
        print(error_prompt)
    return choice


def has_data(working_row):  # 判断当前行的数据区域是否为空，如果为空，返回false，如果不为空，返回true
    for working_cell in working_row:
        if working_cell.value is not None:
            return True
    return False


def check_data(fileNamelist):
    filenames = ['']
    keyvalues = [0]
    index = 0
    checkresult = True
    print("正在进行数据唯一性检测...")
    for filename in fileNamelist:
        workbook_checking = openpyxl.load_workbook(filename)
        worksheet_checking = workbook_checking.active
        for checkcurrentRow in list(worksheet_checking.rows)[ReservedRow:worksheet_checking.max_row]:
            if has_data(checkcurrentRow[ReservedColumn:worksheet_checking.max_column]):
                keyvalue = checkcurrentRow[KeyColumn-1].value
                # print(keyvalue)
                try:
                    result = keyvalues.index(keyvalue)
                    # print(result)
                except ValueError:
                    filenames.append(str(filename))
                    keyvalues.append(keyvalue)
                    index += 1
                    # print(index)
                else:
                    print("发现数据冲突！文档 {} 中的 {} 与文档 {} 中的 {} 同时存在数据！".format(
                        filenames[result], keyvalues[result], filename, keyvalue))
                    checkresult = False
                    # print(checkresult)
        workbook_checking.close()
    if checkresult:
        print("数据唯一性检测完成，没有发现数据冲突。")
    else:
        print("由于存在数据冲突，无法继续进行数据合并。麻烦丽霞再核对一下数据吧，程序不能再继续了")
        exit(1)


def make_file_bydatetime(originalFilename: str):
    """
    根据当前系统日期和时间，以及原有的合并文档名称，生成新的包含合并数据的新的文档名称
    """
    current_time = datetime.datetime.now()
    newPostfix = str(current_time.date()) + '_' + str(
        current_time.hour) + '-' + str(current_time.minute) + '-' + str(current_time.second)
    newFilename = os.path.splitext(originalFilename)[
        0] + newPostfix + os.path.splitext(originalFilename)[-1]
    try:
        copyfile(originalFilename, newFilename)
    except IOError as e:
        print("无法复制生成新的合并文档：{}。程序不能再继续了".format(e))
        exit(1)
    except:
        print("未知错误：{}".format(sys.exec_info()))
        exit(1)
    return newFilename


# Press the green button in the gutter to run the script.
if __name__ == '__main__':

    settings_value = make_settings()
    ReservedRow = settings_value[0]
    ReservedColumn = settings_value[1]
    KeyColumn = settings_value[2]
    display_currentpath_info()
    user_choice = get_choice()
    filenameCombine = fileNameList.pop(
        user_choice - 1)  # 将用于存放合并数据的Excel文档从列表中取出
    check_data(fileNameList)    # 检查剩余文档中是否存在数据冲突（如果存在则提示后退出程序）

    # 根据系统日期和时间以主文档为蓝本复制生成新的合并主文档
    filenameCombine = make_file_bydatetime(filenameCombine)

    wbc = openpyxl.load_workbook(filenameCombine)
    wsc = wbc.active  # 定位到文档中的活动工作表:worksheet of combine
    keycolumnvalue = [0]
    for i in range(ReservedRow+1, wsc.max_row+1):    # 将用于比对的关键字所在列的内容添加到列表keycolumn中
        keycolumnvalue.append(wsc.cell(i, KeyColumn).value)
    total_lines = 0  # 计数器 - 记录总共合并的记录数
    # 依次打开找到的xlsx文档，处理数据并复制到待合并文档中
    for workingfilename in fileNameList:
        wb = openpyxl.load_workbook(workingfilename)
        ws = wb.active
        count_lines = 0  # 计数器，用于记录在当前文档中被合并了几条记录
        for currentRow in list(ws.rows)[ReservedRow:ws.max_row]:
            if has_data(currentRow[ReservedColumn:ws.max_column]):
                search_value = currentRow[KeyColumn-1].value
                try:
                    find_position = keycolumnvalue.index(search_value)
                except ValueError:
                    print("糟糕！在把文档 {0} 中的记录 {1} 合并到主文档中的时候找不到对应的记录".format(
                        workingfilename, search_value))
                    print("麻烦丽霞再核对一下数据吧，程序不能再继续了")
                    exit(1)
                else:
                    find_position += 1
                    targetRow = wsc[find_position]
                    if has_data(targetRow[ReservedColumn:wsc.max_column]):
                        print("糟糕！在把文档 {0} 中的记录 {1} 合并到主文档中的时候发现主文档当中已有相关数据".format(
                            workingfilename, search_value))
                        print("麻烦丽霞再检查一下主文档吧，程序不能再继续了")
                        exit(1)
                    i = ReservedColumn
                    while i < ws.max_column:
                        wsc.cell(row=find_position, column=i +
                                 1).value = currentRow[i].value
                        i += 1
                    count_lines += 1
        print("从 {} 中合并了 {} 条记录！".format(workingfilename, str(count_lines)))
        total_lines += count_lines
        wbc.save(filenameCombine)
        wb.close()
    wbc.close()
    print("一共从 {} 个文档中合并了 {} 条记录。".format(
        str(len(fileNameList)), str(total_lines)))
